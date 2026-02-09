"""
MacroFactor .xlsx export parser.
Handles three export formats:
  - Format A: Quick Export (single sheet)
  - Format B: Detailed Export (two sheets: Calories & Macros, Scale Weight)
  - Format C: Program Settings (goals/targets)
"""

from typing import Optional
from openpyxl import load_workbook
from backend.utils.logger import get_logger
from backend.utils.date_utils import normalize_date

log = get_logger("MACROFACTOR")

# Validation ranges
WEIGHT_MIN, WEIGHT_MAX = 40.0, 200.0
CALORIES_MIN, CALORIES_MAX = 0, 10000
PROTEIN_MAX = 500
CARBS_MAX = 1000
FATS_MAX = 500
FAT_PERCENT_MIN, FAT_PERCENT_MAX = 1.0, 60.0


class MacroFactorParser:
    """Parse MacroFactor .xlsx exports into standardized data structures."""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.format_type = None

    def load(self):
        """Load the workbook and detect format."""
        log.info("load", f"Loading file: {self.file_path}")
        try:
            self.workbook = load_workbook(self.file_path, read_only=True, data_only=True)
            sheet_names = self.workbook.sheetnames
            log.info("load", f"Sheets found: {sheet_names}")

            self.format_type = self._detect_format(sheet_names)
            log.info("load", f"Detected format: {self.format_type}")
            return self
        except Exception as e:
            log.error("load", f"Failed to load file: {type(e).__name__}: {e}")
            raise ValueError(f"Could not open file: {e}")

    def _detect_format(self, sheet_names: list[str]) -> str:
        """Detect which MacroFactor export format this file is."""
        names_lower = [s.lower().strip() for s in sheet_names]

        if any("quick export" in n for n in names_lower):
            return "quick_export"
        elif any("calories" in n and "macros" in n for n in names_lower):
            return "detailed_export"
        elif any("program settings" in n for n in names_lower):
            return "program_settings"
        else:
            raise ValueError(
                f"Unknown MacroFactor format. Expected 'Quick Export', 'Calories & Macros', "
                f"or 'Program Settings' sheet. Found: {sheet_names}"
            )

    def parse(self) -> dict:
        """
        Parse the file and return standardized data.

        Returns:
            {
                "format": str,
                "daily_data": {date_str: {field: value}},  # For formats A & B
                "goals_data": [{"weekday": ..., ...}],      # For format C
                "summary": str,                              # Human-readable summary
            }
        """
        if not self.workbook:
            self.load()

        if self.format_type == "quick_export":
            return self._parse_quick_export()
        elif self.format_type == "detailed_export":
            return self._parse_detailed_export()
        elif self.format_type == "program_settings":
            return self._parse_program_settings()
        else:
            raise ValueError(f"Unknown format type: {self.format_type}")

    # ─── Format A: Quick Export ──────────────────────────────────────

    def _parse_quick_export(self) -> dict:
        """
        Parse Quick Export format.
        Columns: Date | Expenditure | Trend Weight (kg) | Weight (kg) | Steps | Calories (kcal) | Protein (g) | Fat (g) | Carbs (g)
        """
        log.info("parse", "Parsing Quick Export format")
        ws = self._get_sheet_by_pattern("quick export")
        daily_data = {}

        headers = self._read_headers(ws)
        col_map = self._map_columns(headers, {
            "date": "Date",
            "expenditure": "Expenditure",
            "trend weight": "Trend_Weight_kg",
            "weight": "Weight_kg",
            "steps": "Steps_MF",  # MacroFactor steps (we prefer Oura, but store for reference)
            "calories": "Calories",
            "protein": "Protein_g",
            "fat": "Fats_g",
            "carbs": "Carbs_g",
        })

        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            date = normalize_date(row[col_map.get("Date", 0)] if "Date" in col_map else row[0])
            if not date:
                continue

            entry = {}
            self._safe_set(entry, "Expenditure", row, col_map, int_type=True)
            self._safe_set(entry, "Trend_Weight_kg", row, col_map, float_type=True)
            self._safe_set(entry, "Weight_kg", row, col_map, float_type=True)
            self._safe_set(entry, "Calories", row, col_map, int_type=True)
            self._safe_set(entry, "Protein_g", row, col_map, float_type=True)
            self._safe_set(entry, "Fats_g", row, col_map, float_type=True)
            self._safe_set(entry, "Carbs_g", row, col_map, float_type=True)

            # Validate
            entry = self._validate_entry(entry, date)

            if entry:
                daily_data[date] = entry
                row_count += 1

        dates = sorted(daily_data.keys())
        date_range_str = f"{dates[0]}–{dates[-1]}" if dates else "none"
        summary = f"Quick Export: {row_count} days parsed ({date_range_str})"
        log.info("parse", f"SUCCESS — {summary}")

        return {
            "format": "quick_export",
            "daily_data": daily_data,
            "goals_data": None,
            "summary": summary,
        }

    # ─── Format B: Detailed Export ───────────────────────────────────

    def _parse_detailed_export(self) -> dict:
        """
        Parse Detailed Export format (two sheets).
        Sheet 1: Calories & Macros — Date | Calories (kcal) | Protein (g) | Fat (g) | Carbs (g)
        Sheet 2: Scale Weight — Date | Weight (kg) | Fat Percent
        """
        log.info("parse", "Parsing Detailed Export format")
        daily_data = {}

        # Parse Calories & Macros sheet
        ws_macros = self._get_sheet_by_pattern("calories")
        if ws_macros:
            headers = self._read_headers(ws_macros)
            col_map = self._map_columns(headers, {
                "date": "Date",
                "calories": "Calories",
                "protein": "Protein_g",
                "fat": "Fats_g",
                "carbs": "Carbs_g",
            })

            for row in ws_macros.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                date = normalize_date(row[col_map.get("Date", 0)] if "Date" in col_map else row[0])
                if not date:
                    continue

                entry = daily_data.get(date, {})
                self._safe_set(entry, "Calories", row, col_map, int_type=True)
                self._safe_set(entry, "Protein_g", row, col_map, float_type=True)
                self._safe_set(entry, "Fats_g", row, col_map, float_type=True)
                self._safe_set(entry, "Carbs_g", row, col_map, float_type=True)
                daily_data[date] = entry

        # Parse Scale Weight sheet
        ws_weight = self._get_sheet_by_pattern("scale weight")
        if ws_weight:
            headers = self._read_headers(ws_weight)
            col_map = self._map_columns(headers, {
                "date": "Date",
                "weight": "Weight_kg",
                "fat percent": "Fat_Percent",
            })

            for row in ws_weight.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                date = normalize_date(row[col_map.get("Date", 0)] if "Date" in col_map else row[0])
                if not date:
                    continue

                entry = daily_data.get(date, {})
                self._safe_set(entry, "Weight_kg", row, col_map, float_type=True)
                self._safe_set(entry, "Fat_Percent", row, col_map, float_type=True)
                daily_data[date] = entry

        # Validate all entries
        validated = {}
        for date, entry in daily_data.items():
            validated_entry = self._validate_entry(entry, date)
            if validated_entry:
                validated[date] = validated_entry

        dates = sorted(validated.keys())
        date_range_str = f"{dates[0]}–{dates[-1]}" if dates else "none"
        weight_count = sum(1 for d in validated.values() if d.get("Weight_kg"))
        nutrition_count = sum(1 for d in validated.values() if d.get("Calories"))
        summary = f"Detailed Export: {len(validated)} days ({date_range_str}), {nutrition_count} nutrition, {weight_count} weight entries"
        log.info("parse", f"SUCCESS — {summary}")

        return {
            "format": "detailed_export",
            "daily_data": validated,
            "goals_data": None,
            "summary": summary,
        }

    # ─── Format C: Program Settings ─────────────────────────────────

    def _parse_program_settings(self) -> dict:
        """
        Parse Program Settings format (goals/targets per weekday).
        Columns: Program Update Date | Program Weekday | Calories (kcal) | Fat (g) | Protein (g) | Carbs (g) | ...
        Uses the most recent program update for each weekday.
        """
        log.info("parse", "Parsing Program Settings format")
        ws = self._get_sheet_by_pattern("program settings")
        headers = self._read_headers(ws)

        col_map = self._map_columns(headers, {
            "program update date": "Update_Date",
            "program weekday": "Weekday",
            "calories": "Target_Calories",
            "fat": "Target_Fats_g",
            "protein": "Target_Protein_g",
            "carbs": "Target_Carbs_g",
            "expenditure": "Expenditure",
            "weight": "Target_Weight_kg",
        })

        # Collect all entries, keyed by (weekday, update_date)
        entries = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            update_date = normalize_date(
                row[col_map.get("Update_Date", 0)] if "Update_Date" in col_map else row[0]
            )
            weekday_idx = col_map.get("Weekday", 1)
            weekday = row[weekday_idx] if weekday_idx < len(row) else None

            if not update_date or weekday is None:
                continue

            entry = {
                "update_date": update_date,
                "weekday": weekday,
            }

            # Extract target values
            for field in ("Target_Calories", "Target_Protein_g", "Target_Carbs_g", "Target_Fats_g", "Target_Weight_kg"):
                idx = col_map.get(field)
                if idx is not None and idx < len(row) and row[idx] is not None:
                    try:
                        entry[field] = float(row[idx]) if field == "Target_Weight_kg" else int(float(row[idx]))
                    except (ValueError, TypeError):
                        entry[field] = None

            entries.append(entry)

        # Get the most recent program update for each weekday
        # Map weekday names to standard names
        weekday_map = {
            0: "Monday", 1: "Tuesday", 2: "Wednesday", 3: "Thursday",
            4: "Friday", 5: "Saturday", 6: "Sunday",
            "monday": "Monday", "tuesday": "Tuesday", "wednesday": "Wednesday",
            "thursday": "Thursday", "friday": "Friday", "saturday": "Saturday",
            "sunday": "Sunday",
        }

        latest_by_weekday = {}
        for entry in entries:
            weekday = entry["weekday"]
            # Normalize weekday
            if isinstance(weekday, (int, float)):
                weekday_name = weekday_map.get(int(weekday), f"Day_{int(weekday)}")
            else:
                weekday_name = weekday_map.get(str(weekday).lower().strip(), str(weekday))

            if weekday_name not in latest_by_weekday or entry["update_date"] > latest_by_weekday[weekday_name]["update_date"]:
                latest_by_weekday[weekday_name] = entry
                latest_by_weekday[weekday_name]["weekday_name"] = weekday_name

        # Convert to goals format
        goals_data = []
        weekday_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        for day_name in weekday_order:
            if day_name in latest_by_weekday:
                e = latest_by_weekday[day_name]
                goals_data.append({
                    "Weekday": day_name,
                    "Target_Calories": e.get("Target_Calories", ""),
                    "Target_Protein_g": e.get("Target_Protein_g", ""),
                    "Target_Carbs_g": e.get("Target_Carbs_g", ""),
                    "Target_Fats_g": e.get("Target_Fats_g", ""),
                    "Target_Weight_kg": e.get("Target_Weight_kg", ""),
                    "Last_Updated": e.get("update_date", ""),
                })

        # Build summary
        if goals_data:
            sample = goals_data[0]
            summary = (
                f"Program Settings: Updated targets for {len(goals_data)} days. "
                f"Sample (Mon): {sample.get('Target_Calories', '?')}cal | "
                f"{sample.get('Target_Protein_g', '?')}P {sample.get('Target_Carbs_g', '?')}C {sample.get('Target_Fats_g', '?')}F"
            )
        else:
            summary = "Program Settings: No valid goal data found"

        log.info("parse", f"SUCCESS — {summary}")

        return {
            "format": "program_settings",
            "daily_data": None,
            "goals_data": goals_data,
            "summary": summary,
        }

    # ─── Helpers ─────────────────────────────────────────────────────

    def _get_sheet_by_pattern(self, pattern: str):
        """Find a sheet whose name contains the pattern (case-insensitive)."""
        pattern_lower = pattern.lower()
        for name in self.workbook.sheetnames:
            if pattern_lower in name.lower():
                return self.workbook[name]
        return None

    @staticmethod
    def _read_headers(ws) -> list[str]:
        """Read the first row as lowercase header strings."""
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
            headers.append(str(cell).lower().strip() if cell else "")
        return headers

    @staticmethod
    def _map_columns(headers: list[str], mapping: dict[str, str]) -> dict[str, int]:
        """
        Map expected column patterns to their indices.
        mapping: {pattern_substring: field_name}
        Returns: {field_name: column_index}
        """
        col_map = {}
        for pattern, field_name in mapping.items():
            for i, header in enumerate(headers):
                if pattern in header:
                    col_map[field_name] = i
                    break
        return col_map

    @staticmethod
    def _safe_set(entry: dict, field: str, row: tuple, col_map: dict,
                  int_type: bool = False, float_type: bool = False):
        """Safely extract a value from a row and set it in the entry dict."""
        idx = col_map.get(field)
        if idx is None or idx >= len(row):
            return
        value = row[idx]
        if value is None or value == "" or value == "N/A":
            return

        try:
            if int_type:
                entry[field] = int(float(value))
            elif float_type:
                entry[field] = round(float(value), 1)
            else:
                entry[field] = value
        except (ValueError, TypeError):
            pass

    @staticmethod
    def _validate_entry(entry: dict, date: str) -> Optional[dict]:
        """Validate field ranges and log warnings for outliers."""
        if not entry:
            return None

        validated = {}
        for field, value in entry.items():
            if value is None:
                continue

            # Weight validation
            if field == "Weight_kg" and isinstance(value, (int, float)):
                if not (WEIGHT_MIN <= value <= WEIGHT_MAX):
                    log.warning("validate", f"{date}: Weight {value}kg outside range ({WEIGHT_MIN}-{WEIGHT_MAX})")
                    continue

            # Calories validation
            if field == "Calories" and isinstance(value, (int, float)):
                if not (CALORIES_MIN <= value <= CALORIES_MAX):
                    log.warning("validate", f"{date}: Calories {value} outside range ({CALORIES_MIN}-{CALORIES_MAX})")
                    continue

            # Protein validation
            if field == "Protein_g" and isinstance(value, (int, float)):
                if value > PROTEIN_MAX:
                    log.warning("validate", f"{date}: Protein {value}g exceeds {PROTEIN_MAX}g")
                    continue

            # Fat Percent validation
            if field == "Fat_Percent" and isinstance(value, (int, float)):
                if not (FAT_PERCENT_MIN <= value <= FAT_PERCENT_MAX):
                    log.warning("validate", f"{date}: Fat% {value} outside range ({FAT_PERCENT_MIN}-{FAT_PERCENT_MAX})")
                    continue

            validated[field] = value

        return validated if validated else None

    def close(self):
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()
